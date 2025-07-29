import requests
import pandas as pd
import json
import os
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import Rule, CellIsRule
from openpyxl.utils import get_column_letter

# --- Configuration ---
SPECIES_COUNT = 1025 
API_SPECIES_URL = "https://pokeapi.co/api/v2/pokemon-species/"
CACHE_FILENAME = "pokemon_cache.json"

def apply_table_formatting(worksheet):
    """Applies header, banding, and column width formatting to a worksheet."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    banded_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    # Format header
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Format data rows with banding
    for row in range(2, worksheet.max_row + 1):
        if row % 2 == 0:
            for cell in worksheet[row]:
                cell.fill = banded_fill

    # Auto-fit columns
    for col in worksheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) if max_length > 0 else 12
        worksheet.column_dimensions[column].width = adjusted_width

def get_all_pokemon_forms_data():
    """
    Fetches data for all Pokémon forms from PokéAPI or a local cache file.
    """
    if os.path.exists(CACHE_FILENAME):
        print(f"Cache file '{CACHE_FILENAME}' found! Loading data from cache.")
        print("This will be much faster. To force a refresh, delete the cache file.")
        with open(CACHE_FILENAME, 'r', encoding='utf-8') as f:
            all_pokemon_data = json.load(f)
        all_pokemon_data.sort(key=lambda p: p['#'])
        return pd.DataFrame(all_pokemon_data)

    print("No cache file found. Fetching fresh data from PokéAPI.")
    all_pokemon_data = []
    print(f"Fetching all Pokémon forms for {SPECIES_COUNT} species. This will take several minutes...")

    for i in range(1, SPECIES_COUNT + 1):
        try:
            species_res = requests.get(f"{API_SPECIES_URL}{i}")
            species_res.raise_for_status()
            species_data = species_res.json()
            
            species_name = species_data['name'].title()
            print(f"\nFetching Species #{i}: {species_name}")

            for variety in species_data['varieties']:
                pokemon_name = variety['pokemon']['name']
                pokemon_url = variety['pokemon']['url']

                try:
                    print(f"  -> Fetching variant: {pokemon_name}")
                    pokemon_res = requests.get(pokemon_url)
                    pokemon_res.raise_for_status()
                    data = pokemon_res.json()

                    stats = {stat['stat']['name'].replace('-', ' '): stat['base_stat'] for stat in data['stats']}
                    abilities = data['abilities']
                    ability_1 = next((a['ability']['name'] for a in abilities if a['slot'] == 1), None)
                    ability_2 = next((a['ability']['name'] for a in abilities if a['slot'] == 2), None)
                    hidden_ability = next((a['ability']['name'] for a in abilities if a['is_hidden']), None)
                    formatted_name = data['name'].replace('-', ' ').title()
                    
                    pokemon_entry = {
                        '#': data['id'], 'Name': formatted_name,
                        'Type 1': data['types'][0]['type']['name'].title(),
                        'Type 2': data['types'][1]['type']['name'].title() if len(data['types']) > 1 else None,
                        'Ability 1': ability_1.replace('-', ' ').title() if ability_1 else None,
                        'Ability 2': ability_2.replace('-', ' ').title() if ability_2 else None,
                        'Hidden Ability': hidden_ability.replace('-', ' ').title() if hidden_ability else None,
                        'HP': stats.get('hp', 0), 'Attack': stats.get('attack', 0),
                        'Defense': stats.get('defense', 0), 'Sp Atk': stats.get('special attack', 0),
                        'Sp Def': stats.get('special defense', 0), 'Speed': stats.get('speed', 0),
                    }
                    all_pokemon_data.append(pokemon_entry)
                except requests.exceptions.RequestException as e:
                    print(f"    -> ERROR fetching variant {pokemon_name}: {e}")
        except requests.exceptions.RequestException as e:
            print(f"ERROR fetching species with ID {i}: {e}")

    print(f"\nSaving data to cache file '{CACHE_FILENAME}' for future runs...")
    with open(CACHE_FILENAME, 'w', encoding='utf-8') as f:
        json.dump(all_pokemon_data, f, indent=4)
    
    all_pokemon_data.sort(key=lambda p: p['#'])
    return pd.DataFrame(all_pokemon_data)

def create_excel_file(df, filename="Pokemon_Stat_Editor_All_Forms.xlsx"):
    """Creates the Excel workbook with database, editor, and results sheets."""
    if df.empty:
        print("No data was fetched, cannot create Excel file.")
        return

    stat_cols = ['HP', 'Attack', 'Defense', 'Sp Atk', 'Sp Def', 'Speed']
    df['BST'] = df[stat_cols].sum(axis=1)

    all_abilities = pd.concat([df['Ability 1'], df['Ability 2'], df['Hidden Ability']]).dropna().unique()
    all_abilities.sort()
    all_types = pd.concat([df['Type 1'], df['Type 2']]).dropna().unique()
    all_types.sort()

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # --- Sheet 1: Pokedex Database ---
        df.to_excel(writer, sheet_name='Pokedex Database', index=False)
        # --- Sheet 2: Hidden Lists ---
        pd.DataFrame({'Abilities': all_abilities}).to_excel(writer, sheet_name='Lists', index=False, startcol=0)
        pd.DataFrame({'Types': all_types}).to_excel(writer, sheet_name='Lists', index=False, startcol=2)
        # --- Sheet 3: Pokemon Editor ---
        pd.DataFrame().to_excel(writer, sheet_name='Pokemon Editor', index=False)
        # --- Sheet 4: Edited Pokedex ---
        pd.DataFrame(columns=df.columns).to_excel(writer, sheet_name='Edited Pokedex', index=False)
        
        workbook = writer.book
        
        # --- Format Pokedex Database sheet ---
        apply_table_formatting(workbook['Pokedex Database'])
        
        # --- Format Edited Pokedex sheet (header only for now) ---
        apply_table_formatting(workbook['Edited Pokedex'])

        # --- Setup Pokemon Editor Sheet ---
        editor_sheet = workbook['Pokemon Editor']
        db_sheet_name = 'Pokedex Database'
        
        # Styles for Editor
        header_style = Font(bold=True, size=12)
        category_style = Font(bold=True)
        
        headers = [
            ("B2", "Select Pokémon:", None), ("A4", "Category", category_style), 
            ("B4", "ORIGINAL", header_style), ("D4", "EDITED", header_style),
            ("A5", "Ability 1", category_style), ("A6", "Ability 2", category_style), 
            ("A7", "Hidden Ability", category_style), ("A8", "Type 1", category_style), 
            ("A9", "Type 2", category_style), ("A11", "HP", category_style), 
            ("A12", "Attack", category_style), ("A13", "Defense", category_style), 
            ("A14", "Sp. Atk", category_style), ("A15", "Sp. Def", category_style), 
            ("A16", "Speed", category_style), ("A18", "BASE STAT TOTAL", category_style)
        ]
        for cell_ref, value, style in headers: 
            editor_sheet[cell_ref] = value
            if style: editor_sheet[cell_ref].font = style
        
        editor_sheet["D18"] = "=+SUM(D11:D16)"
        editor_sheet["D18"].font = category_style

        # Data Validation
        dv_pokemon = DataValidation(type="list", formula1=f"'{db_sheet_name}'!$B$2:$B${len(df) + 1}")
        editor_sheet.add_data_validation(dv_pokemon)
        dv_pokemon.add('C2')
        dv_abilities = DataValidation(type="list", formula1=f"'Lists'!$A$2:$A${len(all_abilities) + 1}")
        editor_sheet.add_data_validation(dv_abilities)
        dv_abilities.add('D5:D7')
        dv_types = DataValidation(type="list", formula1=f"'Lists'!$C$2:$C${len(all_types) + 1}")
        editor_sheet.add_data_validation(dv_types)
        dv_types.add('D8:D9')
        
        # VLOOKUP Formulas
        lookup_map = {
            'B5': 4, 'B6': 5, 'B7': 6, 'B8': 2, 'B9': 3, 'B11': 7, 'B12': 8, 
            'B13': 9, 'B14': 10, 'B15': 11, 'B16': 12, 'B18': 13
        }
        for cell, col_idx in lookup_map.items():
            editor_sheet[cell] = f'=IFERROR(VLOOKUP($C$2, \'{db_sheet_name}\'!$B:$N, {col_idx}, FALSE), "")'

        # Conditional Formatting
        blue_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid') 
        lightblue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        lightgreen_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        stat_rules = [
            CellIsRule(operator='greaterThanOrEqual', formula=[150], fill=blue_fill),
            CellIsRule(operator='greaterThanOrEqual', formula=[120], fill=lightblue_fill),
            CellIsRule(operator='greaterThanOrEqual', formula=[100], fill=green_fill),
            CellIsRule(operator='greaterThanOrEqual', formula=[90], fill=lightgreen_fill),
            CellIsRule(operator='greaterThanOrEqual', formula=[60], fill=yellow_fill),
            CellIsRule(operator='lessThan', formula=[60], fill=orange_fill)
        ]
        for r in ["B11:B16", "D11:D16"]:
            for rule in stat_rules: editor_sheet.conditional_formatting.add(r, rule)
        
        type_colors = {
            'Normal': 'A8A878', 'Fire': 'F08030', 'Water': '6890F0', 'Electric': 'F8D030',
            'Grass': '78C850', 'Ice': '98D8D8', 'Fighting': 'C03028', 'Poison': 'A040A0',
            'Ground': 'E0C068', 'Flying': 'A890F0', 'Psychic': 'F85888', 'Bug': 'A8B820',
            'Rock': 'B8A038', 'Ghost': '705898', 'Dragon': '7038F8', 'Dark': '705848',
            'Steel': 'B8B8D0', 'Fairy': 'EE99AC'
        }
        white_font = Font(color="FFFFFF")
        for r in ["B8:B9", "D8:D9"]:
            for type_name, color_hex in type_colors.items():
                fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                font = white_font if type_name in ['Fighting', 'Poison', 'Ghost', 'Dark', 'Dragon', 'Fire'] else None
                rule = CellIsRule(operator='equal', formula=[f'"{type_name}"'], fill=fill, font=font)
                editor_sheet.conditional_formatting.add(r, rule)

        workbook['Lists'].sheet_state = 'hidden'

        print(f"\nExcel file '{filename}' created successfully!")
        print("All sheets are now visually formatted.")

if __name__ == "__main__":
    pokemon_df = get_all_pokemon_forms_data()
    create_excel_file(pokemon_df)
