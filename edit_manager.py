import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os
import sys
import subprocess

EXCEL_FILENAME = "Pokemon_Stat_Editor_All_Forms.xlsx"

def apply_banding_to_row(worksheet, row_number):
    """Applies the alternating row color to a specific row."""
    banded_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for cell in worksheet[row_number]:
        cell.fill = PatternFill(fill_type=None)
    if row_number % 2 == 0:
        for cell in worksheet[row_number]:
            cell.fill = banded_fill

def update_editor_view_and_get_selection(workbook):
    """
    Updates the editor's dropdown and populates the 'ORIGINAL' column
    based on the current selection. This replaces the need for VLOOKUPs.
    """
    # --- Create a unified view of all Pokémon ---
    db_df = pd.read_excel(EXCEL_FILENAME, sheet_name="Pokedex Database")
    edited_df = pd.read_excel(EXCEL_FILENAME, sheet_name="Edited Pokedex")
    
    combined_df = pd.concat([edited_df, db_df]).drop_duplicates(subset=['Name'], keep='first')
    combined_df = combined_df.sort_values(by='#').reset_index(drop=True)

    # --- Update the dropdown list with all Pokémon names ---
    editor_sheet = workbook["Pokemon Editor"]
    lists_sheet = workbook["Lists"]
    
    all_names = combined_df['Name'].tolist()
    # Write the unified list to a new column in the 'Lists' sheet (e.g., column E)
    # Clear old list first
    for row in lists_sheet['E']:
        row.value = None
    lists_sheet['E1'] = "All Pokemon Names"
    for i, name in enumerate(all_names, 2):
        lists_sheet[f'E{i}'] = name

    # Update DataValidation for the main selector
    dv_pokemon = DataValidation(type="list", formula1=f"'Lists'!$E$2:$E${len(all_names) + 1}")
    editor_sheet.add_data_validation(dv_pokemon)
    dv_pokemon.add('C2')

    # --- Populate the 'ORIGINAL' column based on the selection ---
    selected_name = editor_sheet['C2'].value
    if not selected_name:
        return None, None # No selection made

    try:
        pokemon_data = combined_df[combined_df['Name'] == selected_name].iloc[0]
    except IndexError:
        print(f"Warning: '{selected_name}' not found in any database. Clearing fields.")
        pokemon_data = {} # Create empty dict to clear fields

    # Directly write data to the 'ORIGINAL' column
    editor_sheet['B5'] = pokemon_data.get('Ability 1', '')
    editor_sheet['B6'] = pokemon_data.get('Ability 2', '')
    editor_sheet['B7'] = pokemon_data.get('Hidden Ability', '')
    editor_sheet['B8'] = pokemon_data.get('Type 1', '')
    editor_sheet['B9'] = pokemon_data.get('Type 2', '')
    editor_sheet['B11'] = pokemon_data.get('HP', '')
    editor_sheet['B12'] = pokemon_data.get('Attack', '')
    editor_sheet['B13'] = pokemon_data.get('Defense', '')
    editor_sheet['B14'] = pokemon_data.get('Sp Atk', '')
    editor_sheet['B15'] = pokemon_data.get('Sp Def', '')
    editor_sheet['B16'] = pokemon_data.get('Speed', '')
    editor_sheet['B18'] = pokemon_data.get('BST', '')
    
    return selected_name, db_df # Return the name and the original db dataframe

def commit_changes():
    """
    Reads data from the 'Pokemon Editor', then updates an existing entry
    or appends a new one to the 'Edited Pokedex' sheet.
    """
    try:
        with open(EXCEL_FILENAME, 'a+b') as f:
            pass
    except PermissionError:
        print("--- ACTION REQUIRED ---")
        print(f"Error: The file '{EXCEL_FILENAME}' is currently open. Please close it and run this script again.")
        return
    except FileNotFoundError:
        print(f"Error: '{EXCEL_FILENAME}' not found. Please run 'pokedata.py' first.")
        return

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILENAME)
        
        # Update dropdowns and get current selection
        selected_name, db_df = update_editor_view_and_get_selection(workbook)
        
        if not selected_name:
            print("No Pokémon selected in the editor. Dropdown lists have been updated.")
            workbook.save(EXCEL_FILENAME)
            print("File saved. Please select a Pokémon and run the script again.")
            return

        print(f"Processing edits for: {selected_name}")
        editor_sheet = workbook["Pokemon Editor"]

        # --- Conditional Filling: Use original value if edited cell is blank ---
        original_values = {
            'Ability 1': editor_sheet['B5'].value, 'Ability 2': editor_sheet['B6'].value,
            'Hidden Ability': editor_sheet['B7'].value, 'Type 1': editor_sheet['B8'].value,
            'Type 2': editor_sheet['B9'].value, 'HP': editor_sheet['B11'].value,
            'Attack': editor_sheet['B12'].value, 'Defense': editor_sheet['B13'].value,
            'Sp Atk': editor_sheet['B14'].value, 'Sp Def': editor_sheet['B15'].value,
            'Speed': editor_sheet['B16'].value,
        }
        
        edited_values = {
            'Ability 1': editor_sheet['D5'].value, 'Ability 2': editor_sheet['D6'].value,
            'Hidden Ability': editor_sheet['D7'].value, 'Type 1': editor_sheet['D8'].value,
            'Type 2': editor_sheet['D9'].value, 'HP': editor_sheet['D11'].value,
            'Attack': editor_sheet['D12'].value, 'Defense': editor_sheet['D13'].value,
            'Sp Atk': editor_sheet['D14'].value, 'Sp Def': editor_sheet['D15'].value,
            'Speed': editor_sheet['D16'].value,
        }

        final_data = {key: edited_values[key] if edited_values[key] is not None else original_values[key] for key in original_values}
        
        # Get original Pokédex number and construct the final row
        original_pokemon_row = db_df[db_df['Name'] == selected_name].iloc[0]
        final_data['#'] = original_pokemon_row['#']
        final_data['Name'] = selected_name
        
        stat_cols = ['HP', 'Attack', 'Defense', 'Sp Atk', 'Sp Def', 'Speed']
        numeric_stats = [pd.to_numeric(final_data[stat], errors='coerce') for stat in stat_cols]
        final_data['BST'] = sum(s for s in numeric_stats if pd.notna(s))
        
        final_row_data = [final_data.get(col, '') for col in db_df.columns]
        
        # --- Update or Append to 'Edited Pokedex' ---
        edited_sheet = workbook["Edited Pokedex"]
        target_row = None
        for row in range(2, edited_sheet.max_row + 1):
            if edited_sheet.cell(row=row, column=2).value == selected_name:
                target_row = row
                break
        
        if target_row:
            print(f"Found existing entry for '{selected_name}'. Overwriting...")
            for col_idx, cell_value in enumerate(final_row_data, 1):
                edited_sheet.cell(row=target_row, column=col_idx).value = cell_value
            apply_banding_to_row(edited_sheet, target_row)
            message = f"Successfully updated '{selected_name}'!"
        else:
            print(f"No entry found for '{selected_name}'. Appending new row...")
            edited_sheet.append(final_row_data)
            apply_banding_to_row(edited_sheet, edited_sheet.max_row)
            message = f"Successfully added '{selected_name}'!"

        workbook.save(EXCEL_FILENAME)
        print(message)
        
        print(f"Attempting to re-open '{EXCEL_FILENAME}'...")
        if sys.platform == "win32": os.startfile(EXCEL_FILENAME)
        elif sys.platform == "darwin": subprocess.run(["open", EXCEL_FILENAME], check=True)
        else: subprocess.run(["xdg-open", EXCEL_FILENAME], check=True)

    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    commit_changes()
