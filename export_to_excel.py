import pandas as pd
import json
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def apply_table_formatting(worksheet):
    """Applies header, banding, and column width formatting to a worksheet."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    banded_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    if worksheet.max_row == 0: # Handle empty sheet
        return

    # Format header
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Format data rows with banding
    for row_index in range(2, worksheet.max_row + 1):
        if row_index % 2 == 0:
            for cell in worksheet[row_index]:
                cell.fill = banded_fill

    # Auto-fit columns
    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) if max_length > 0 else 12
        worksheet.column_dimensions[column_letter].width = adjusted_width

def export_json_to_excel():
    """
    Reads a JSON file containing Pokémon data and exports it to a
    nicely formatted Excel file.
    """
    # --- Get input from the user ---
    input_filename = input("Enter the name of the input JSON file (e.g., custom_pokemon.json): ")
    output_filename = input("Enter the desired name for the output Excel file (e.g., custom_pokedex.xlsx): ")

    # Ensure the output filename has the correct extension
    if not output_filename.endswith('.xlsx'):
        output_filename += '.xlsx'

    try:
        # --- Read and process the data ---
        # Use pandas to easily read the JSON into a DataFrame
        df = pd.read_json(input_filename)
        
        # Ensure columns are in the desired order
        desired_columns = [
            '#', 'Name', 'Type 1', 'Type 2', 'Ability 1', 'Ability 2', 'Hidden Ability',
            'HP', 'Attack', 'Defense', 'Sp Atk', 'Sp Def', 'Speed', 'BST'
        ]
        # Filter to only include columns that exist in the DataFrame
        existing_columns = [col for col in desired_columns if col in df.columns]
        df = df[existing_columns]

        print(f"\nSuccessfully read {len(df)} Pokémon from '{input_filename}'.")
        print(f"Exporting to '{output_filename}'...")

        # --- Write to Excel and apply formatting ---
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Custom Pokedex', index=False)
            
            # Access the workbook and worksheet to apply formatting
            workbook = writer.book
            worksheet = workbook['Custom Pokedex']
            
            apply_table_formatting(worksheet)

        print("\nExcel file created successfully!")

    except FileNotFoundError:
        print(f"\nError: The file '{input_filename}' was not found. Please make sure it's in the same folder.")
    except Exception as e:
        print(f"\nAn unexpected error occurred: {e}")

# --- Main Execution ---
if __name__ == "__main__":
    export_json_to_excel()
